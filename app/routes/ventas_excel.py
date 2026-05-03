from flask import Blueprint, request, send_file
from app.database import get_connection
from app.utils.response import success, error
from app.utils.email import enviar_alerta_stock
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO

bp_excel = Blueprint("ventas_excel", __name__)

# ─────────────────────────────────────────
# GET /ventas/plantilla → Descargar plantilla Excel
# ─────────────────────────────────────────
@bp_excel.route("/plantilla", methods=["GET"])
def descargar_plantilla():
    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Traer clientes existentes
        cur.execute("""
            SELECT id_cliente, folio, nombre, apellido_paterno
            FROM clientes
            WHERE folio != 'PUB-001'
            ORDER BY id_cliente ASC
        """)
        clientes = cur.fetchall()

        # Traer último folio de cliente
        cur.execute("""
            SELECT folio FROM clientes
            WHERE folio != 'PUB-001'
            ORDER BY id_cliente DESC
            LIMIT 1
        """)
        ultimo_cliente = cur.fetchone()
        ultimo_folio = ultimo_cliente["folio"] if ultimo_cliente else "No hay clientes registrados"

        # Traer inventarios disponibles
        cur.execute("""
            SELECT
                p.id_producto,
                p.descripcion,
                a.id_almacen,
                a.nombre AS nombre_almacen,
                i.stock
            FROM inventarios i
            LEFT JOIN productos p ON p.id_producto = i.id_producto
            LEFT JOIN almacenes a ON a.id_almacen = i.id_almacen
            WHERE i.stock > 0
            ORDER BY p.id_producto ASC, a.id_almacen ASC
        """)
        inventarios = cur.fetchall()

        # Traer estados y municipios
        cur.execute("""
            SELECT
                e.id_estado,
                e.nombre AS nombre_estado,
                m.id_municipio,
                m.nombre AS nombre_municipio
            FROM estados e
            LEFT JOIN municipios m ON m.id_estado = e.id_estado
            ORDER BY e.nombre ASC, m.nombre ASC
        """)
        estados_municipios = cur.fetchall()

        wb = Workbook()

        # ── Estilos ──
        header_font = Font(bold=True, color="FFFFFF")
        header_fill_blue = PatternFill("solid", start_color="2F5496")
        header_fill_green = PatternFill("solid", start_color="375623")
        header_fill_orange = PatternFill("solid", start_color="C55A11")
        header_fill_gray = PatternFill("solid", start_color="595959")
        header_fill_purple = PatternFill("solid", start_color="6B2D8B")
        header_alignment = Alignment(horizontal="center", vertical="center")
        ejemplo_font = Font(italic=True, color="808080")
        nota_font = Font(italic=True, color="595959")

        # ── Hoja oculta: Listas para dropdowns ──
        ws_listas = wb.create_sheet(title="_listas")
        ws_listas.sheet_state = "hidden"

        # Columna A: combinaciones producto-almacen
        combinaciones = []
        for inv in inventarios:
            combinacion = f"{inv['descripcion']} - {inv['nombre_almacen']}"
            combinaciones.append(combinacion)
            ws_listas.append([combinacion])

        # Columna B: clientes
        for i, cliente in enumerate(clientes, start=1):
            etiqueta = f"{cliente['folio']} - {cliente['nombre']} {cliente['apellido_paterno']}"
            ws_listas.cell(row=i, column=2, value=etiqueta)

        total_combinaciones = len(combinaciones)
        total_clientes = len(clientes)

        # ── Hoja 1: Ventas ──
        ws_ventas = wb.active
        ws_ventas.title = "Ventas"

        enc_ventas = ["numero_venta", "folio", "precio_venta_final", "cliente", "id_estado", "id_municipio"]
        for col, enc in enumerate(enc_ventas, start=1):
            cell = ws_ventas.cell(row=1, column=col, value=enc)
            cell.font = header_font
            cell.fill = header_fill_blue
            cell.alignment = header_alignment

        ejemplo_cliente = f"{clientes[0]['folio']} - {clientes[0]['nombre']} {clientes[0]['apellido_paterno']}" if clientes else "CLI-001 - Juan Pérez"
        ejemplos_ventas = [1, "VTA-010", "", ejemplo_cliente, 30, 3625]
        for col, val in enumerate(ejemplos_ventas, start=1):
            cell = ws_ventas.cell(row=2, column=col, value=val)
            cell.font = ejemplo_font

        # Dropdown de clientes en columna D
        if total_clientes > 0:
            dv_clientes = DataValidation(
                type="list",
                formula1=f"_listas!$B$1:$B${total_clientes}",
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Cliente inválido",
                error="Selecciona un cliente de la lista."
            )
            ws_ventas.add_data_validation(dv_clientes)
            dv_clientes.sqref = "D2:D1000"

        ws_ventas.column_dimensions["A"].width = 15
        ws_ventas.column_dimensions["B"].width = 15
        ws_ventas.column_dimensions["C"].width = 22
        ws_ventas.column_dimensions["D"].width = 35
        ws_ventas.column_dimensions["E"].width = 12
        ws_ventas.column_dimensions["F"].width = 15

        ws_ventas.cell(row=4, column=1, value="NOTAS:").font = Font(bold=True)
        notas_ventas = [
            "• 'numero_venta' es un identificador temporal solo para este archivo (1, 2, 3...). No se guarda en el sistema.",
            "• 'precio_venta_final' es opcional, se calculará automáticamente sumando (cantidad * precio_venta) de cada producto.",
            "• 'cliente' es obligatorio. Selecciona de la lista desplegable o registra uno nuevo en la hoja 'Clientes'.",
            "• Si el cliente no aparece en la lista, regístralo en la sección 'REGISTRAR CLIENTES NUEVOS' de la hoja 'Clientes' y escribe su folio manualmente aquí.",
            "• 'id_estado' e 'id_municipio' son opcionales. Consulta la hoja 'Catálogo Estados y Municipios' para encontrar los IDs. Usa Ctrl+F para buscar.",
            "• Borra la fila de ejemplo (fila 2) antes de subir el archivo."
        ]
        for i, nota in enumerate(notas_ventas, start=5):
            cell = ws_ventas.cell(row=i, column=1, value=nota)
            cell.font = nota_font

        # ── Hoja 2: Detalle de ventas ──
        ws_detalle = wb.create_sheet(title="Detalle Ventas")

        enc_detalle = ["numero_venta", "producto_almacen", "cantidad_vendida"]
        for col, enc in enumerate(enc_detalle, start=1):
            cell = ws_detalle.cell(row=1, column=col, value=enc)
            cell.font = header_font
            cell.fill = header_fill_green
            cell.alignment = header_alignment

        ejemplo_inv = combinaciones[0] if combinaciones else "Panel solar - Almacen 1"
        ejemplos_detalle = [1, ejemplo_inv, 2]
        for col, val in enumerate(ejemplos_detalle, start=1):
            cell = ws_detalle.cell(row=2, column=col, value=val)
            cell.font = ejemplo_font

        # Dropdown de producto-almacen en columna B
        if total_combinaciones > 0:
            dv_inv = DataValidation(
                type="list",
                formula1=f"_listas!$A$1:$A${total_combinaciones}",
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Combinación inválida",
                error="Selecciona una combinación de producto y almacén de la lista."
            )
            ws_detalle.add_data_validation(dv_inv)
            dv_inv.sqref = "B2:B1000"

        ws_detalle.column_dimensions["A"].width = 15
        ws_detalle.column_dimensions["B"].width = 45
        ws_detalle.column_dimensions["C"].width = 18

        ws_detalle.cell(row=4, column=1, value="NOTAS:").font = Font(bold=True)
        notas_detalle = [
            "• 'numero_venta' debe coincidir con el de la hoja 'Ventas' para relacionar cada producto con su venta.",
            "• 'producto_almacen' usa la lista desplegable para seleccionar la combinación correcta.",
            "• 'cantidad_vendida' debe ser mayor a 0.",
            "• El precio de venta se tomará automáticamente del precio registrado de cada producto en el sistema.",
            "• Borra la fila de ejemplo (fila 2) antes de subir el archivo."
        ]
        for i, nota in enumerate(notas_detalle, start=5):
            cell = ws_detalle.cell(row=i, column=1, value=nota)
            cell.font = nota_font

        # ── Hoja 3: Clientes ──
        ws_clientes = wb.create_sheet(title="Clientes")

        ws_clientes.cell(row=1, column=1, value="CLIENTES EXISTENTES").font = Font(bold=True, color="FFFFFF")
        ws_clientes.cell(row=1, column=1).fill = header_fill_blue

        enc_clientes_existentes = ["id_cliente", "folio", "nombre", "apellido_paterno"]
        for col, enc in enumerate(enc_clientes_existentes, start=1):
            cell = ws_clientes.cell(row=2, column=col, value=enc)
            cell.font = header_font
            cell.fill = header_fill_gray
            cell.alignment = header_alignment

        for row, cliente in enumerate(clientes, start=3):
            ws_clientes.cell(row=row, column=1, value=cliente["id_cliente"])
            ws_clientes.cell(row=row, column=2, value=cliente["folio"])
            ws_clientes.cell(row=row, column=3, value=cliente["nombre"])
            ws_clientes.cell(row=row, column=4, value=cliente["apellido_paterno"])

        fila_separador = len(clientes) + 4
        ws_clientes.cell(row=fila_separador, column=1, value="REGISTRAR CLIENTES NUEVOS").font = Font(bold=True, color="FFFFFF")
        ws_clientes.cell(row=fila_separador, column=1).fill = header_fill_orange
        ws_clientes.cell(row=fila_separador + 1, column=1, value=f"Último folio registrado: {ultimo_folio}. El siguiente folio debería continuar esa secuencia.").font = nota_font

        enc_clientes_nuevos = ["folio", "nombre", "apellido_paterno", "apellido_materno", "telefono", "email"]
        for col, enc in enumerate(enc_clientes_nuevos, start=1):
            cell = ws_clientes.cell(row=fila_separador + 2, column=col, value=enc)
            cell.font = header_font
            cell.fill = header_fill_orange
            cell.alignment = header_alignment

        ws_clientes.column_dimensions["A"].width = 15
        ws_clientes.column_dimensions["B"].width = 20
        ws_clientes.column_dimensions["C"].width = 20
        ws_clientes.column_dimensions["D"].width = 20
        ws_clientes.column_dimensions["E"].width = 15
        ws_clientes.column_dimensions["F"].width = 25

        # ── Hoja 4: Catálogo de inventarios ──
        ws_inv = wb.create_sheet(title="Catálogo Inventarios")

        ws_inv.cell(row=1, column=1, value="INVENTARIO DISPONIBLE").font = Font(bold=True, color="FFFFFF")
        ws_inv.cell(row=1, column=1).fill = header_fill_green

        enc_inv = ["descripcion", "nombre_almacen", "stock_disponible", "precio_venta"]
        for col, enc in enumerate(enc_inv, start=1):
            cell = ws_inv.cell(row=2, column=col, value=enc)
            cell.font = header_font
            cell.fill = header_fill_gray
            cell.alignment = header_alignment

        for row, inv in enumerate(inventarios, start=3):
            # Traer precio del producto
            cur.execute("SELECT precio FROM productos WHERE id_producto = %s", (inv["id_producto"],))
            precio = cur.fetchone()
            ws_inv.cell(row=row, column=1, value=inv["descripcion"])
            ws_inv.cell(row=row, column=2, value=inv["nombre_almacen"])
            ws_inv.cell(row=row, column=3, value=inv["stock"])
            ws_inv.cell(row=row, column=4, value=precio["precio"] if precio else None)

        ws_inv.column_dimensions["A"].width = 35
        ws_inv.column_dimensions["B"].width = 25
        ws_inv.column_dimensions["C"].width = 18
        ws_inv.column_dimensions["D"].width = 15

        # ── Hoja 5: Catálogo Estados y Municipios ──
        ws_em = wb.create_sheet(title="Catálogo Estados y Municipios")

        ws_em.cell(row=1, column=1, value="ESTADOS Y MUNICIPIOS").font = Font(bold=True, color="FFFFFF")
        ws_em.cell(row=1, column=1).fill = header_fill_purple

        ws_em.cell(row=2, column=1, value="Usa Ctrl+F para buscar tu estado o municipio y copia el ID correspondiente a la hoja de Ventas.").font = nota_font

        enc_em = ["id_estado", "nombre_estado", "id_municipio", "nombre_municipio"]
        for col, enc in enumerate(enc_em, start=1):
            cell = ws_em.cell(row=3, column=col, value=enc)
            cell.font = header_font
            cell.fill = header_fill_gray
            cell.alignment = header_alignment

        for row, em in enumerate(estados_municipios, start=4):
            ws_em.cell(row=row, column=1, value=em["id_estado"])
            ws_em.cell(row=row, column=2, value=em["nombre_estado"])
            ws_em.cell(row=row, column=3, value=em["id_municipio"])
            ws_em.cell(row=row, column=4, value=em["nombre_municipio"])

        ws_em.column_dimensions["A"].width = 12
        ws_em.column_dimensions["B"].width = 25
        ws_em.column_dimensions["C"].width = 15
        ws_em.column_dimensions["D"].width = 30

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="plantilla_ventas.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return error(message=str(e), status=500)
    finally:
        if conn:
            cur.close()
            conn.close()


# ─────────────────────────────────────────
# POST /ventas/carga-masiva → Cargar ventas desde Excel
# ─────────────────────────────────────────
@bp_excel.route("/carga-masiva", methods=["POST"])
def carga_masiva():
    conn = None
    try:
        if "archivo" not in request.files:
            return error(message="No se encontró el archivo en la solicitud", status=400)

        archivo = request.files["archivo"]
        if archivo.filename == "":
            return error(message="No se seleccionó ningún archivo", status=400)
        if not archivo.filename.endswith(".xlsx"):
            return error(message="El archivo debe ser de tipo .xlsx", status=400)

        conn = get_connection()
        cur = conn.cursor()

        wb = load_workbook(archivo, data_only=True)

        for hoja in ["Ventas", "Detalle Ventas", "Clientes"]:
            if hoja not in wb.sheetnames:
                return error(message=f"No se encontró la hoja '{hoja}' en el archivo", status=400)

        ws_ventas = wb["Ventas"]
        ws_detalle = wb["Detalle Ventas"]
        ws_clientes = wb["Clientes"]

        errores = []

        # ── Paso 1: Leer y validar clientes nuevos ──
        clientes_nuevos = []
        folios_clientes_nuevos = []

        fila_inicio_nuevos = None
        for row in ws_clientes.iter_rows():
            for cell in row:
                if cell.value == "REGISTRAR CLIENTES NUEVOS":
                    fila_inicio_nuevos = cell.row + 2
                    break
            if fila_inicio_nuevos:
                break

        if fila_inicio_nuevos:
            for num_fila, fila in enumerate(ws_clientes.iter_rows(min_row=fila_inicio_nuevos + 1, values_only=True), start=fila_inicio_nuevos + 1):
                if not any(fila):
                    continue

                folio, nombre, apellido_paterno, apellido_materno, telefono, email = (list(fila) + [None] * 6)[:6]

                if not folio:
                    errores.append({"hoja": "Clientes", "fila": num_fila, "motivo": "El campo 'folio' es obligatorio para clientes nuevos"})
                    continue
                if not nombre:
                    errores.append({"hoja": "Clientes", "fila": num_fila, "motivo": "El campo 'nombre' es obligatorio para clientes nuevos"})
                    continue
                if not apellido_paterno:
                    errores.append({"hoja": "Clientes", "fila": num_fila, "motivo": "El campo 'apellido_paterno' es obligatorio para clientes nuevos"})
                    continue

                cur.execute("SELECT folio FROM clientes WHERE folio = %s", (str(folio),))
                if cur.fetchone() is not None:
                    errores.append({"hoja": "Clientes", "fila": num_fila, "motivo": f"El folio '{folio}' ya existe en la base de datos"})
                    continue

                if str(folio) in folios_clientes_nuevos:
                    errores.append({"hoja": "Clientes", "fila": num_fila, "motivo": f"El folio '{folio}' está duplicado en el archivo"})
                    continue

                folios_clientes_nuevos.append(str(folio))
                clientes_nuevos.append({
                    "folio": str(folio),
                    "nombre": nombre,
                    "apellido_paterno": apellido_paterno,
                    "apellido_materno": apellido_materno,
                    "telefono": str(telefono) if telefono else None,
                    "email": str(email) if email else None,
                    "fila": num_fila
                })

        # ── Paso 2: Leer y validar ventas ──
        ventas_dict = {}

        for num_fila, fila in enumerate(ws_ventas.iter_rows(min_row=2, values_only=True), start=2):
            if not any(fila):
                continue
            if str(fila[0] or "").startswith("•") or str(fila[0] or "").strip() in ["NOTAS:", ""]:
                continue

            numero_venta, folio, precio_venta_final, cliente_raw, id_estado, id_municipio = (list(fila) + [None] * 6)[:6]

            if numero_venta is None:
                errores.append({"hoja": "Ventas", "fila": num_fila, "motivo": "El campo 'numero_venta' es obligatorio"})
                continue
            if not cliente_raw:
                errores.append({"hoja": "Ventas", "fila": num_fila, "motivo": f"El campo 'cliente' es obligatorio en la venta {numero_venta}"})
                continue
            if int(numero_venta) in ventas_dict:
                errores.append({"hoja": "Ventas", "fila": num_fila, "motivo": f"El numero_venta {numero_venta} está duplicado en el archivo"})
                continue

            # Extraer folio del cliente desde "CLI-001 - Juan Pérez"
            folio_cliente = str(cliente_raw).split(" - ")[0].strip()

            cur.execute("SELECT id_cliente FROM clientes WHERE folio = %s", (folio_cliente,))
            cliente_bd = cur.fetchone()

            if cliente_bd is None and folio_cliente not in folios_clientes_nuevos:
                errores.append({
                    "hoja": "Ventas",
                    "fila": num_fila,
                    "motivo": f"El cliente con folio '{folio_cliente}' no existe. Regístralo en la sección 'REGISTRAR CLIENTES NUEVOS' de la hoja 'Clientes'"
                })
                continue

            ventas_dict[int(numero_venta)] = {
                "folio": str(folio) if folio else None,
                "precio_venta_final": precio_venta_final,  # Puede ser None, se calculará después
                "folio_cliente": folio_cliente,
                "id_estado": int(id_estado) if id_estado else None,
                "id_municipio": int(id_municipio) if id_municipio else None,
                "fila": num_fila
            }

        # ── Paso 3: Leer y validar detalle ──
        detalle_dict = {}

        for num_fila, fila in enumerate(ws_detalle.iter_rows(min_row=2, values_only=True), start=2):
            if not any(fila):
                continue
            if str(fila[0] or "").startswith("•") or str(fila[0] or "").strip() in ["NOTAS:", ""]:
                continue

            numero_venta, producto_almacen_raw, cantidad_vendida = (list(fila) + [None] * 3)[:3]

            if numero_venta is None:
                errores.append({"hoja": "Detalle Ventas", "fila": num_fila, "motivo": "El campo 'numero_venta' es obligatorio"})
                continue

            numero_venta = int(numero_venta)

            if numero_venta not in ventas_dict:
                errores.append({"hoja": "Detalle Ventas", "fila": num_fila, "motivo": f"El numero_venta {numero_venta} no existe en la hoja 'Ventas'"})
                continue
            if not producto_almacen_raw:
                errores.append({"hoja": "Detalle Ventas", "fila": num_fila, "motivo": f"El campo 'producto_almacen' es obligatorio en la fila {num_fila}"})
                continue
            if cantidad_vendida is None:
                errores.append({"hoja": "Detalle Ventas", "fila": num_fila, "motivo": f"El campo 'cantidad_vendida' es obligatorio en la fila {num_fila}"})
                continue
            if cantidad_vendida <= 0:
                errores.append({"hoja": "Detalle Ventas", "fila": num_fila, "motivo": f"La cantidad_vendida debe ser mayor a 0 en la fila {num_fila}"})
                continue

            partes = str(producto_almacen_raw).split(" - ")
            if len(partes) < 2:
                errores.append({"hoja": "Detalle Ventas", "fila": num_fila, "motivo": f"El formato de 'producto_almacen' es inválido en la fila {num_fila}"})
                continue

            descripcion_producto = partes[0].strip()
            nombre_almacen = partes[1].strip()

            # Buscar id_producto, id_almacen, precio y stock
            cur.execute("""
                SELECT
                    i.id_producto,
                    i.id_almacen,
                    i.stock,
                    p.precio
                FROM inventarios i
                LEFT JOIN productos p ON p.id_producto = i.id_producto
                LEFT JOIN almacenes a ON a.id_almacen = i.id_almacen
                WHERE p.descripcion = %s AND a.nombre = %s AND i.stock > 0
            """, (descripcion_producto, nombre_almacen))
            inventario = cur.fetchone()

            if inventario is None:
                errores.append({"hoja": "Detalle Ventas", "fila": num_fila, "motivo": f"No se encontró inventario disponible para '{producto_almacen_raw}'"})
                continue
            if cantidad_vendida > inventario["stock"]:
                errores.append({"hoja": "Detalle Ventas", "fila": num_fila, "motivo": f"Stock insuficiente para '{producto_almacen_raw}'. Stock disponible: {inventario['stock']}"})
                continue

            if numero_venta not in detalle_dict:
                detalle_dict[numero_venta] = []

            detalle_dict[numero_venta].append({
                "id_producto": inventario["id_producto"],
                "id_almacen": inventario["id_almacen"],
                "cantidad_vendida": int(cantidad_vendida),
                "precio_venta": inventario["precio"] or 0
            })

        # Verificar que todas las ventas tienen al menos un producto
        for numero_venta, venta in ventas_dict.items():
            if numero_venta not in detalle_dict:
                errores.append({"hoja": "Detalle Ventas", "fila": venta["fila"], "motivo": f"La venta {numero_venta} no tiene ningún producto en la hoja 'Detalle Ventas'"})

        if errores:
            return error(
                message=f"Se encontraron {len(errores)} error(es) en el archivo. No se registró ninguna venta.",
                status=400,
                data={"errores": errores}
            )

        if not ventas_dict:
            return error(message="El archivo no contiene ventas para registrar", status=400)

        # ── Paso 4: Insertar clientes nuevos ──
        folios_a_id = {}

        folios_usados = set(v["folio_cliente"] for v in ventas_dict.values())
        for folio in folios_usados:
            if folio not in folios_clientes_nuevos:
                cur.execute("SELECT id_cliente FROM clientes WHERE folio = %s", (folio,))
                resultado = cur.fetchone()
                if resultado:
                    folios_a_id[folio] = resultado["id_cliente"]

        for cliente in clientes_nuevos:
            cur.execute("""
                INSERT INTO clientes (folio, nombre, apellido_paterno, apellido_materno, telefono, email)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id_cliente
            """, (
                cliente["folio"],
                cliente["nombre"],
                cliente["apellido_paterno"],
                cliente["apellido_materno"],
                cliente["telefono"],
                cliente["email"],
            ))
            nuevo_id = cur.fetchone()["id_cliente"]
            folios_a_id[cliente["folio"]] = nuevo_id

        # ── Paso 5: Insertar ventas y detalle ──
        ventas_creadas = 0
        for numero_venta, venta in ventas_dict.items():
            id_cliente = folios_a_id.get(venta["folio_cliente"])
            items_detalle = detalle_dict[numero_venta]

            # Calcular precio_venta_final si no se mandó
            precio_venta_final = venta["precio_venta_final"]
            if not precio_venta_final:
                precio_venta_final = sum(item["cantidad_vendida"] * item["precio_venta"] for item in items_detalle)

            cur.execute("""
                INSERT INTO ventas (folio, precio_venta_final, id_estado, id_municipio, id_cliente)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id_venta
            """, (
                venta["folio"],
                precio_venta_final,
                venta["id_estado"],
                venta["id_municipio"],
                id_cliente,
            ))
            nuevo_id_venta = cur.fetchone()["id_venta"]

            for item in items_detalle:
                cur.execute("""
                    INSERT INTO detalle_venta (cantidad_vendida, precio_venta, id_venta, id_producto)
                    VALUES (%s, %s, %s, %s)
                """, (
                    item["cantidad_vendida"],
                    item["precio_venta"],
                    nuevo_id_venta,
                    item["id_producto"],
                ))

                cur.execute("""
                    INSERT INTO movimientos_inventario (tipo, cantidad, id_venta, id_producto, id_almacen)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    False,
                    item["cantidad_vendida"],
                    nuevo_id_venta,
                    item["id_producto"],
                    item["id_almacen"],
                ))

                cur.execute("""
                    UPDATE inventarios SET
                        stock = stock - %s
                    WHERE id_producto = %s AND id_almacen = %s
                """, (
                    item["cantidad_vendida"],
                    item["id_producto"],
                    item["id_almacen"],
                ))

                # Verificar stock mínimo
                cur.execute("""
                    SELECT
                        i.stock,
                        i.min_stock,
                        p.descripcion,
                        a.nombre AS nombre_almacen
                    FROM inventarios i
                    LEFT JOIN productos p ON p.id_producto = i.id_producto
                    LEFT JOIN almacenes a ON a.id_almacen = i.id_almacen
                    WHERE i.id_producto = %s AND i.id_almacen = %s
                """, (item["id_producto"], item["id_almacen"]))
                inventario_actualizado = cur.fetchone()

                if inventario_actualizado["stock"] <= inventario_actualizado["min_stock"]:
                    enviar_alerta_stock(
                        descripcion_producto=inventario_actualizado["descripcion"],
                        nombre_almacen=inventario_actualizado["nombre_almacen"],
                        stock_actual=inventario_actualizado["stock"],
                        min_stock=inventario_actualizado["min_stock"]
                    )

            ventas_creadas += 1

        conn.commit()
        return success(
            data={
                "ventas_creadas": ventas_creadas,
                "clientes_registrados": len(clientes_nuevos)
            },
            message=f"Se registraron {ventas_creadas} venta(s) y {len(clientes_nuevos)} cliente(s) nuevo(s) correctamente",
            status=201
        )

    except Exception as e:
        if conn:
            conn.rollback()
        return error(message=str(e), status=500)
    finally:
        if conn:
            cur.close()
            conn.close()