from flask import Blueprint, request, send_file
from app.database import get_connection
from app.utils.response import success, error
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from app.utils.jwt import verificar_token, requiere_admin

bp_excel = Blueprint("productos_excel", __name__)

# ─────────────────────────────────────────
# GET /productos/plantilla/<id_cat> → Descargar plantilla Excel
# ─────────────────────────────────────────
@bp_excel.route("/plantilla/<int:id_cat>", methods=["GET"])
def descargar_plantilla(id_cat):
    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Verificar que la categoría existe
        cur.execute("SELECT id_cat, nombre FROM categorias WHERE id_cat = %s", (id_cat,))
        categoria = cur.fetchone()
        if categoria is None:
            return error(message="La categoría seleccionada no existe", status=404)

        wb = Workbook()

        # ── Hoja principal: Productos ──
        ws_productos = wb.active
        ws_productos.title = "Productos"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="2F5496")
        header_alignment = Alignment(horizontal="center", vertical="center")
        ejemplo_font = Font(italic=True, color="808080")
        nota_font = Font(italic=True, color="595959")

        encabezados = ["folio", "descripcion", "costo", "moneda", "margenes"]
        for col, encabezado in enumerate(encabezados, start=1):
            cell = ws_productos.cell(row=1, column=col, value=encabezado)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Fila de ejemplo
        ejemplos = ["PROD-001", "Descripción del producto", 50.00, "MXN", "10,20,30"]
        for col, valor in enumerate(ejemplos, start=1):
            cell = ws_productos.cell(row=2, column=col, value=valor)
            cell.font = ejemplo_font

        ws_productos.column_dimensions["A"].width = 15
        ws_productos.column_dimensions["B"].width = 35
        ws_productos.column_dimensions["C"].width = 12
        ws_productos.column_dimensions["D"].width = 10
        ws_productos.column_dimensions["E"].width = 25

        ws_productos.cell(row=4, column=1, value="NOTAS:").font = Font(bold=True)
        notas = [
            f"• Todos los productos se registrarán en la categoría: {categoria['nombre']}",
            "• Los campos 'folio', 'costo', 'moneda' y 'margenes' son obligatorios.",
            "• 'moneda' debe ser 'MXN' para pesos o 'USD' para dólares.",
            "• En 'margenes' escribe los porcentajes separados por coma. Ejemplo: 10,20,30",
            "• Borra la fila de ejemplo (fila 2) antes de subir el archivo."
        ]
        for i, nota in enumerate(notas, start=5):
            cell = ws_productos.cell(row=i, column=1, value=nota)
            cell.font = nota_font

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"plantilla_productos_{categoria['nombre'].replace(' ', '_')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return error(message=str(e), status=500)
    finally:
        if conn:
            cur.close()
            conn.close()


# ─────────────────────────────────────────
# POST /productos/carga-masiva/<id_cat> → Cargar productos desde Excel
# ─────────────────────────────────────────
@bp_excel.route("/carga-masiva/<int:id_cat>", methods=["POST"])
def carga_masiva(id_cat):
    conn = None
    try:
        if "archivo" not in request.files:
            return error(message="No se encontró el archivo en la solicitud", status=400)

        archivo = request.files["archivo"]
        if archivo.filename == "":
            return error(message="No se seleccionó ningún archivo", status=400)
        if not archivo.filename.endswith(".xlsx"):
            return error(message="El archivo debe ser de tipo .xlsx", status=400)

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT id_cat, nombre FROM categorias WHERE id_cat = %s", (id_cat,))
        categoria = cur.fetchone()
        if categoria is None:
            return error(message="La categoría seleccionada no existe", status=404)

        wb = load_workbook(archivo, data_only=True)
        ws = wb.active

        errores = []
        productos_a_crear = []

        for num_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(fila):
                continue
            if str(fila[0] or "").startswith("•") or str(fila[0] or "").strip() in ["NOTAS:", ""]:
                continue

            folio, descripcion, costo, moneda, margenes_raw = (list(fila) + [None] * 5)[:5]

            # Validar campos obligatorios
            if not folio:
                errores.append({"fila": num_fila, "motivo": "El campo 'folio' es obligatorio"})
                continue
            if not costo:
                errores.append({"fila": num_fila, "motivo": "El campo 'costo' es obligatorio"})
                continue
            if not moneda:
                errores.append({"fila": num_fila, "motivo": "El campo 'moneda' es obligatorio"})
                continue
            if str(moneda).upper() not in ["MXN", "USD"]:
                errores.append({"fila": num_fila, "motivo": f"El campo 'moneda' debe ser 'MXN' o 'USD'"})
                continue
            if not margenes_raw:
                errores.append({"fila": num_fila, "motivo": "El campo 'margenes' es obligatorio"})
                continue

            # Verificar folio duplicado en BD
            cur.execute("SELECT folio FROM productos WHERE folio = %s", (str(folio),))
            if cur.fetchone() is not None:
                errores.append({"fila": num_fila, "motivo": f"El folio '{folio}' ya existe en la base de datos"})
                continue

            # Verificar folio duplicado dentro del mismo archivo
            folios_en_archivo = [p["folio"] for p in productos_a_crear]
            if str(folio) in folios_en_archivo:
                errores.append({"fila": num_fila, "motivo": f"El folio '{folio}' está duplicado en el archivo"})
                continue

            # Parsear márgenes
            try:
                margenes = [int(m.strip()) for m in str(margenes_raw).split(",") if m.strip()]
                if len(margenes) == 0:
                    raise ValueError
            except ValueError:
                errores.append({"fila": num_fila, "motivo": "El campo 'margenes' contiene valores inválidos, deben ser números enteros separados por coma"})
                continue

            productos_a_crear.append({
                "folio": str(folio),
                "descripcion": descripcion,
                "costo": costo,
                "moneda": str(moneda).upper(),
                "margenes": margenes,
                "fila": num_fila
            })

        if errores:
            return error(
                message=f"Se encontraron {len(errores)} error(es) en el archivo. No se creó ningún producto.",
                status=400,
                data={"errores": errores}
            )

        if not productos_a_crear:
            return error(message="El archivo no contiene productos para registrar", status=400)

        ids_creados = []
        for producto in productos_a_crear:
            cur.execute("""
                INSERT INTO productos (folio, descripcion, costo, moneda)
                VALUES (%s, %s, %s, %s)
                RETURNING id_producto
            """, (
                producto["folio"],
                producto["descripcion"],
                producto["costo"],
                producto["moneda"],
            ))
            nuevo_id = cur.fetchone()["id_producto"]

            # Asignar categoría
            cur.execute("""
                INSERT INTO producto_categoria (id_prod, id_cat)
                VALUES (%s, %s)
            """, (nuevo_id, id_cat))

            # Calcular e insertar precios por margen
            for margen in producto["margenes"]:
                precio_margen = producto["costo"] / (1 - (margen / 100))
                cur.execute("""
                    INSERT INTO productos_precios (id_producto, margen, precio_margen)
                    VALUES (%s, %s, %s)
                """, (nuevo_id, margen, round(precio_margen, 2)))

            ids_creados.append(nuevo_id)

        conn.commit()
        return success(
            data={
                "productos_creados": len(ids_creados),
                "ids": ids_creados
            },
            message=f"Se crearon {len(ids_creados)} producto(s) correctamente",
            status=201
        )

    except Exception as e:
        if conn:
            conn.rollback()
        return error(message=str(e), status=500)
    finally:
        if conn:
            cur.close()
            conn.close()
